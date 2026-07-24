import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

import supabase_client

# ------------------------------------------------------------
# In-memory cache of the geography config, loaded from Supabase.
# Call refresh_geo_cache() after editing data in the admin dashboard
# so a running app picks up changes without a redeploy.
# ------------------------------------------------------------

state_to_region = {}
region_to_rsn = {}
state_to_field_office = {}
chapter_lookup = {}          # (field_office, region) -> chapter, plus ("__default__", region) -> chapter
special_overrides = {}       # field_office -> {"region": ..., "rsn": ..., "state": ...}
state_splits = pd.DataFrame(columns=["Zip Code", "Main City", "County", "STATE", "ASSIGNED TO"])
mapping = {}                 # keyword -> program (same name/shape irfas_cleaner.py already expects)

states_of_interest = ['CA', 'TX', 'FL', 'VA', 'MO']

zero_value_df = pd.read_csv("zero_values_2026_starter.csv")


def refresh_geo_cache():
    """Reload all geography config tables from Supabase into memory."""
    global state_to_region, region_to_rsn, state_to_field_office
    global chapter_lookup, special_overrides, state_splits, mapping

    state_to_region = {
        row["state"]: row["region"]
        for row in supabase_client.fetch_all("geo_state_to_region")
    }

    region_to_rsn = {
        row["region"]: row["rsn"]
        for row in supabase_client.fetch_all("geo_region_to_rsn")
    }

    state_to_field_office = {
        row["state"]: row["field_office"]
        for row in supabase_client.fetch_all("geo_state_to_field_office")
    }

    chapter_lookup = {
        (row["field_office"], row["region"]): row["chapter"]
        for row in supabase_client.fetch_all("geo_chapter_map")
    }

    special_overrides = {
        row["when_field_office"]: {
            "region": row.get("set_region"),
            "rsn": row.get("set_rsn"),
            "state": row.get("set_state"),
        }
        for row in supabase_client.fetch_all("geo_special_overrides")
        if row.get("active")
    }

    splits_rows = supabase_client.fetch_all("geo_state_splits_cities")
    if splits_rows:
        splits_df = pd.DataFrame(splits_rows)
        splits_df = splits_df.rename(columns={
            "zip_code": "Zip Code",
            "main_city": "Main City",
            "county": "County",
            "state": "STATE",
            "assigned_to": "ASSIGNED TO",
        })
        splits_df["Zip Code"] = (
            splits_df["Zip Code"].fillna("").astype(str).str.split(".").str[0].str.zfill(5)
        )
        state_splits = splits_df
    else:
        state_splits = pd.DataFrame(columns=["Zip Code", "Main City", "County", "STATE", "ASSIGNED TO"])

    mapping = {
        row["keyword"]: row["program"]
        for row in supabase_client.fetch_all("geo_keyword_to_program")
    }


# Load once at import time. If Supabase isn't configured yet (e.g. local
# dev without env vars set), fall back silently and leave the caches
# empty rather than crashing app startup.
try:
    refresh_geo_cache()
except Exception as error:
    print(f"[geography_maps] Could not load geo config from Supabase at startup: {error}")


def assign_chapter(row):
    region = row['Region']
    office = row['Field Office']

    chapter = chapter_lookup.get((office, region))
    if chapter:
        return chapter

    default_chapter = chapter_lookup.get(("__default__", region))
    if default_chapter:
        return default_chapter

    return region  # fallback for any other region


def assign_field_office(row, zip_there):
    if row['State'] in states_of_interest:
        city = row['City']
        state = row['State']
        zip_code = row['Zipcode']

        if isinstance(city, str):
            city = city.title().strip()
        else:
            city = 'Unknown'

        city_match = state_splits[
            (state_splits['Main City'] == city) & (state_splits['STATE'] == state)
        ]

        if not city_match.empty:
            field_office = city_match["ASSIGNED TO"].values[0]
        elif city == 'St Louis':
            field_office = 'St. Louis Office'
        elif city == 'St Petersburg':
            field_office = 'Miami Office'
        elif zip_there and zip_code != "00000":
            zip_match = state_splits[
                (state_splits['Zip Code'] == zip_code) & (state_splits['STATE'] == state)
            ]
            if not zip_match.empty:
                field_office = zip_match["ASSIGNED TO"].values[0]
            else:
                field_office = _default_office_for_state(state, city, zip_code)
        else:
            field_office = _default_office_for_state(state, city, zip_code)

        override = special_overrides.get(field_office)
        if override:
            if override.get("region"):
                row['Region'] = override["region"]
            if override.get("rsn") is not None:
                row['Region Number'] = override["rsn"]
            if override.get("state"):
                row['State'] = override["state"]

        row['Field Office'] = field_office

    return row


def _default_office_for_state(state, city, zip_code):
    state_defaults = {
        'TX': 'Dallas Office',
        'FL': 'Miami Office',
        'CA': 'Los Angeles Office',
        'VA': 'Alexandria Office',
        'MO': 'St. Louis Office'
    }
    field_office = state_defaults.get(state, 'Unknown')
    print(f"No reference for {city}, {state} {zip_code}\nAutomatically assigned to {field_office}\n")
    return field_office


def hq_states(row):
    state = row["State"]
    if (state not in state_to_region.keys()) and (state not in states_of_interest):
        row['State'] = 'HQ'
    return row


def _add_date_columns(df):
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")

    df["Year"] = pd.Series(pd.NA, index=df.index, dtype="Int64")
    df["Month Number"] = pd.Series(pd.NA, index=df.index, dtype="Int64")
    df["Month Name"] = ""
    df["Quarter"] = ""

    valid_date_rows = df["Date"].notna()

    df.loc[valid_date_rows, "Year"] = (
        df.loc[valid_date_rows, "Date"].dt.year.astype("Int64")
    )

    df.loc[valid_date_rows, "Month Number"] = (
        df.loc[valid_date_rows, "Date"].dt.month.astype("Int64")
    )

    df.loc[valid_date_rows, "Month Name"] = (
        df.loc[valid_date_rows, "Date"].dt.month_name()
    )

    quarter_number = (
        ((df.loc[valid_date_rows, "Month Number"] - 1) // 3) + 1
    )

    df.loc[valid_date_rows, "Quarter"] = (
        "Q" + quarter_number.astype(str)
    )

    return df


def save_excel_light(df, output_path):
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Cleaned Volunteers")

    df_to_save = df.copy()
    df_to_save = df_to_save.astype(object)
    df_to_save = df_to_save.where(pd.notna(df_to_save), None)

    for row in dataframe_to_rows(df_to_save, index=False, header=True):
        worksheet.append(row)

    workbook.save(output_path)


def format_currency_column(output_path, currency_column):
    workbook = load_workbook(output_path)
    worksheet = workbook.active

    headers = [cell.value for cell in worksheet[1]]

    currency_column_number = None

    for index, header in enumerate(headers, start=1):
        if str(header).strip().upper() == currency_column.strip().upper():
            currency_column_number = index
            break

    if currency_column_number is None:
        workbook.save(output_path)
        return

    for row_number in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_number, column=currency_column_number)

        value = cell.value

        if value is None or str(value).strip() == "":
            continue

        if isinstance(value, (int, float)):
            numeric_value = value
        else:
            cleaned_value = (
                str(value)
                .strip()
                .replace("$", "")
                .replace(",", "")
            )

            if cleaned_value.startswith("(") and cleaned_value.endswith(")"):
                cleaned_value = "-" + cleaned_value[1:-1]

            try:
                numeric_value = float(cleaned_value)
            except ValueError:
                continue

        cell.value = numeric_value
        cell.number_format = '$#,##0.00'

    workbook.save(output_path)
