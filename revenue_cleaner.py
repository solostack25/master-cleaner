from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import geography_maps
import column_finder as cf

FIRST_REVENUE_HEADER = "Contact: Mailing State/Province ↓"

def normalize_header_value(value):
    if pd.isna(value):
        return ""

    value = str(value).strip()

    # Replace weird spacing/newlines with one normal space
    value = re.sub(r"\s+", " ", value)

    return value


def save_revenue_workbook(cleaned_revenue_df, total_revenue_df, community_revenue_df, output_path):
    output_path = Path(output_path)

    workbook = Workbook(write_only=True)

    cleaned_sheet = workbook.create_sheet("Cleaned Revenue")
    total_sheet = workbook.create_sheet("Total Revenue")
    community_sheet = workbook.create_sheet("Community Revenue")

    cleaned_df_to_save = cleaned_revenue_df.copy()
    cleaned_df_to_save = cleaned_df_to_save.astype(object)
    cleaned_df_to_save = cleaned_df_to_save.where(pd.notna(cleaned_df_to_save), None)

    for row in dataframe_to_rows(cleaned_df_to_save, index=False, header=True):
        cleaned_sheet.append(row)

    total_df_to_save = total_revenue_df.copy()
    total_df_to_save = total_df_to_save.astype(object)
    total_df_to_save = total_df_to_save.where(pd.notna(total_df_to_save), None)

    for row in dataframe_to_rows(total_df_to_save, index=False, header=True):
        total_sheet.append(row)

    community_df_to_save = community_revenue_df.copy()
    community_df_to_save = community_df_to_save.astype(object)
    community_df_to_save = community_df_to_save.where(pd.notna(community_df_to_save), None)

    for row in dataframe_to_rows(community_df_to_save, index=False, header=True):
        community_sheet.append(row)


    workbook.save(output_path)

def clean_revenue_dataframe(input_path):
    input_path = Path(input_path)

    anchor_keyword = normalize_header_value(FIRST_REVENUE_HEADER).lower()
    anchor_keyword = anchor_keyword.replace("↑", "").replace("↓", "").strip()

    df = cf.clean_export_range(input_path, (anchor_keyword,), max_blank_run=2)

    return df

def create_total_revenue_summary(df):
    zero_values_df = geography_maps.zero_value_df

    # gets all of the years in the full detailed row data
    years_in_data = df["YEAR"].dropna().unique()

    years_in_data = sorted(years_in_data)

    zero_value_dataframes = []

    for year in years_in_data:
        year = int(year)

        year_zero_values_df = zero_values_df.copy()

        year_zero_values_df["YEAR"] = year

        if "DATE" in year_zero_values_df.columns:
            year_zero_values_df["DATE"] = pd.to_datetime(
                {
                    "year": year_zero_values_df["YEAR"],
                    "month": year_zero_values_df["MONTH NUMBER"],
                    "day": 1
                }
            )

        zero_value_dataframes.append(year_zero_values_df)

    total_revenue_df = pd.concat(
        zero_value_dataframes,
        ignore_index=True
    )

    revenue_summary = df.groupby(
        [
            "FIELD OFFICE",
            "YEAR",
            "MONTH NUMBER"
        ],
        as_index=False,
        sort=False
    )["PAYMENT AMOUNT"].sum()

    total_revenue_df = total_revenue_df.merge(
        revenue_summary,
        how="left",
        on=[
            "FIELD OFFICE",
            "YEAR",
            "MONTH NUMBER"
        ]
    )

    total_revenue_df["TOTAL REVENUE"] = total_revenue_df["PAYMENT AMOUNT"].fillna(0)

    total_revenue_df = total_revenue_df.drop(
        columns=["PAYMENT AMOUNT"],
        errors="ignore"
    )

    return total_revenue_df


def clean_multiple_revenue_imports(input_paths, output_path):
    output_path = Path(output_path)

    cleaned_dataframes = []

    for input_path in input_paths:
        cleaned_df = clean_revenue_dataframe(input_path)
        cleaned_dataframes.append(cleaned_df)

    if len(cleaned_dataframes) == 0:
        raise ValueError("No revenue files were provided.")

    df = pd.concat(
        cleaned_dataframes,
        ignore_index=True
    )

    # changes the name of some columns

    df = df.rename(columns={'Contact: Mailing State/Province ↓': 'State', 'Contact: Mailing City': 'City',
                            'Contact: Mailing Zip/Postal Code': 'Zipcode', 'Payment Date': 'Date'})

    # Forward Fill the State column
    df['State'] = df['State'].ffill()

    # Changes ZIP Code column to string, fixes leading zeroes, and takes only the first 5 digits
    df["Zipcode"] = df["Zipcode"].astype(str).str.extract(r'(\d{1,5})')[0].str.zfill(5)

    # Makes sure the State column is completely Capitalized
    df["State"] = df["State"].str.upper()

    df = df.apply(geography_maps.hq_states, axis=1)

    df["Region"] = df["State"].map(geography_maps.state_to_region)

    # Assign RSN
    df["Region Number"] = df["Region"].map(geography_maps.region_to_rsn)

    # Assign Field Office
    df["Field Office"] = df["State"].map(geography_maps.state_to_field_office)

    df = df.apply(
        lambda row: geography_maps.assign_field_office(row, True),
        axis=1
    )

    df["Chapter"] = df.apply(geography_maps.assign_chapter, axis=1)
    df['Region'] = df['Region'].str.upper()
    df = geography_maps._add_date_columns(df)

    # Changes the order of the columns
    df = df[["Year", "Quarter", "Month Name", "Month Number", "Date", "Region", "Chapter", "State",
            "Field Office", "Region Number", "City", "Zipcode", "Payment Number", "Primary Contact: Full Name",
             "Payment Amount", "Donation Type", "Primary Campaign Source: Campaign Name", "Opportunity Record Type"]]

    df.columns = df.columns.str.upper()

    # Gives you the Total Revenue Grouped by Month and Field Office
    total_revenue_df = create_total_revenue_summary(df)

    # Selects all rows where the donation is not a grant
    community_revenue_df = df[df["OPPORTUNITY RECORD TYPE"] != "Grant"].reset_index(drop=True)

    # Gives you the Community Revenue Grouped by Month and Field Office
    community_revenue_summary_df = create_total_revenue_summary(community_revenue_df)

    community_revenue_summary_df = community_revenue_summary_df.rename(columns={"TOTAL REVENUE": "COMMUNITY REVENUE"})

    save_revenue_workbook(df, total_revenue_df, community_revenue_summary_df, output_path)