from pathlib import Path
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import geography_maps
from datetime import date
import supabase_client


CAMPAIGN_NAME_TO_PSN = {
    "back 2 school": 1,
    "back2school": 1,
    "disaster relief": 2,
    "disaster relief services": 2,
    "general": 3,
    "grants": 3,             # generic/unspecified campaign -> General
    "capacity building": 3,  # not one of the 9 tracked programs -> General
    "health services": 4,
    "hunger prevention": 5,
    "fate": 6,
    "muslim family services": 7,
    "refugee services": 8,
    "refugee services & community empowerment": 8,
    "transitional housing": 9,
}


def campaign_name_to_psn(campaign_name):
    if pd.isna(campaign_name):
        return 3  # blank Campaign Name -> General

    key = str(campaign_name).strip().lower()
    return CAMPAIGN_NAME_TO_PSN.get(key, 3)  # unrecognized -> General


_grants_goals_cache = None


def refresh_grants_goals_cache():
    global _grants_goals_cache
    _grants_goals_cache = None


def _load_grants_goals():
    global _grants_goals_cache
    if _grants_goals_cache is None:
        rows = supabase_client.fetch_all("grants_field_office_goals")
        _grants_goals_cache = {
            (row["field_office"], int(row["year"])): row["initial_goal"]
            for row in rows
        }
    return _grants_goals_cache


def lookup_goal_amount(field_office, year):
    goals = _load_grants_goals()
    if pd.isna(year):
        return pd.NA
    return goals.get((field_office, int(year)), pd.NA)


def build_grants_detail_worksheet(grants_df):
    """Builds the detailed (non-aggregated) Grants worksheet, matching the
    reference RS-PBI-style layout: REGION, RSN, CHAPTER, ... PSN, QUARTER.
    Expects grants_df with geography + date columns already computed, but
    BEFORE the column names get uppercased for the summary/subtraction step.
    """
    df = grants_df.copy()

    df["PSN"] = df["Campaign Name"].apply(campaign_name_to_psn)

    df["GOAL AMOUNT"] = df.apply(
        lambda row: lookup_goal_amount(row["Field Office"], row["Year"]),
        axis=1,
    )

    df = df.rename(columns={
        "Date": "Payment Date",
        "Region": "REGION",
        "Region Number": "RSN",
        "Chapter": "CHAPTER",
        "Field Office": "FILED OFFICE",  # matches the reference file's spelling exactly
        "Donation Name": "DONATION NAME",
        "Quarter": "QUARTER",
    })

    # State is only used internally for geography assignment -- not part
    # of the reference output, so it's dropped here (Contact State  ↓ is
    # the column the report actually displays).
    df = df.drop(columns=["State"], errors="ignore")

    column_order = [
        "REGION", "RSN", "CHAPTER", "Contact State ↓", "FILED OFFICE",
        "Donation Record Type", "Payment Date", "Contact: Full Name",
        "Payment Amount", "Campaign Name", "GOAL AMOUNT", "City",
        "Payment Method", "ICNA Donations Types", "Notes", "Lead Source",
        "Donation ID", "Comments", "Donation Ref", "Donations Snapshot ID",
        "DONATION NAME", "Year", "Month Name", "Month Number", "PSN",
        "QUARTER",
    ]

    existing_columns = [c for c in column_order if c in df.columns]
    df = df[existing_columns]

    return df


FIRST_CURRENT_REVENUE_HEADER = "Contact: Mailing State/Province ↓"
LAST_CURRENT_REVENUE_HEADER = "Sum of Payment Amount"

MONTH_NAME_TO_NUMBER = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

def parse_current_revenue_filename(original_filename):
    filename = Path(original_filename).stem.lower()

    filename = filename.replace("-", "_")
    filename = filename.replace(" ", "_")

    year_match = re.search(r"(20\d{2})", filename)

    if year_match is None:
        raise ValueError(
            f"Could not find a year in this file name: {original_filename}. "
            "Use a name like total_revenue_2026_01.xlsx or grants_2026.xlsx."
        )

    year = int(year_match.group(1))

    if "grant" in filename:
        return {
            "file_type": "grants",
            "year": year,
            "month_number": None
        }
    else:
        # if "total" in filename and "revenue" in filename:
        month_number = None

        # Example: total_revenue_2026_01
        year_month_match = re.search(r"(20\d{2})_+(0?[1-9]|1[0-2])", filename)

        if year_month_match is not None:
            month_number = int(year_month_match.group(2))

        # Example: total_revenue_01_2026
        if month_number is None:
            month_year_match = re.search(r"(0?[1-9]|1[0-2])_+(20\d{2})", filename)

            if month_year_match is not None:
                month_number = int(month_year_match.group(1))

        # Example: total_revenue_january_2026
        if month_number is None:
            filename_parts = filename.split("_")

            for part in filename_parts:
                if part in MONTH_NAME_TO_NUMBER:
                    month_number = MONTH_NAME_TO_NUMBER[part]
                    break

        if month_number is None:
            raise ValueError(
                f"Could not find a month in this total revenue file name: {original_filename}. "
                "Use a name like total_revenue_2026_01.xlsx."
            )

        return {
            "file_type": "total_revenue",
            "year": year,
            "month_number": month_number
        }

def normalize_header_value(value):
    if pd.isna(value):
        return ""

    value = str(value).strip()

    # Converts weird spacing/newlines into one normal space.
    value = re.sub(r"\s+", " ", value)

    return value


def find_current_revenue_header_bounds(df):
    first_header_clean = normalize_header_value(FIRST_CURRENT_REVENUE_HEADER)
    last_header_clean = normalize_header_value(LAST_CURRENT_REVENUE_HEADER)

    for row_number, row in df.iterrows():
        first_column_number = None
        last_column_number = None

        for column_number, value in enumerate(row):
            value_clean = normalize_header_value(value)

            if value_clean == first_header_clean:
                first_column_number = column_number

            if value_clean == last_header_clean:
                last_column_number = column_number

        if first_column_number is not None and last_column_number is not None:
            if last_column_number < first_column_number:
                raise ValueError(
                    "Current revenue header row was found, but Sum of Payment Amount appears before the first column."
                )

            return row_number, first_column_number, last_column_number

    raise ValueError(
        "Could not find the current revenue header row. "
        f"Expected '{FIRST_CURRENT_REVENUE_HEADER}' through '{LAST_CURRENT_REVENUE_HEADER}'."
    )

def clean_current_revenue_dataframe(input_path):
    input_path = Path(input_path)

    df = pd.read_excel(
        input_path,
        sheet_name=0,
        header=None
    )

    header_row_number, first_column_number, last_column_number = find_current_revenue_header_bounds(df)

    # Keep only:
    # - the header row and everything below it
    # - the first expected column through Sum of Payment Amount
    df = df.iloc[
        header_row_number:,
        first_column_number:last_column_number + 1
    ]

    df = df.reset_index(drop=True)

    # Make first row the headers.
    df.columns = [
        normalize_header_value(value)
        for value in df.iloc[0]
    ]

    # Delete the header row from the data.
    df = df.iloc[1:]
    df = df.reset_index(drop=True)

    # Treat blank-looking cells as empty.
    df = df.replace(r"^\s*$", pd.NA, regex=True)

    # Delete fully empty rows.
    df = df.dropna(axis=0, how="all")
    df = df.reset_index(drop=True)

    return df

def find_grants_header_bounds(df):
    FIRST_CURRENT_REVENUE_HEADER = "Contact State  ↓"
    LAST_CURRENT_REVENUE_HEADER = "Zipcode"
    first_header_clean = normalize_header_value(FIRST_CURRENT_REVENUE_HEADER)
    last_header_clean = normalize_header_value(LAST_CURRENT_REVENUE_HEADER)

    for row_number, row in df.iterrows():
        first_column_number = None
        last_column_number = None

        for column_number, value in enumerate(row):
            value_clean = normalize_header_value(value)

            if value_clean == first_header_clean:
                first_column_number = column_number

            if value_clean == last_header_clean:
                last_column_number = column_number

        if first_column_number is not None and last_column_number is not None:
            if last_column_number < first_column_number:
                raise ValueError("Grants row was found, but Zipcode appears before the first column.")

            return row_number, first_column_number, last_column_number

    raise ValueError("Could not find the current grants header row. " f"Expected '{FIRST_CURRENT_REVENUE_HEADER}' "
                     f"through '{LAST_CURRENT_REVENUE_HEADER}'.")


def clean_grants_dataframe(input_path):
    input_path = Path(input_path)

    df = pd.read_excel(
        input_path,
        sheet_name=0,
        header=None
    )

    header_row_number, first_column_number, last_column_number = find_grants_header_bounds(df)

    # Keep only:
    # - the header row and everything below it
    # - the first expected column through Zipcode
    df = df.iloc[
        header_row_number:,
        first_column_number:last_column_number + 1
    ]

    df = df.reset_index(drop=True)

    # Make first row the headers.
    df.columns = [
        normalize_header_value(value)
        for value in df.iloc[0]
    ]

    # Delete the header row from the data.
    df = df.iloc[1:]
    df = df.reset_index(drop=True)

    # Treat blank-looking cells as empty.
    df = df.replace(r"^\s*$", pd.NA, regex=True)

    # Delete fully empty rows.
    df = df.dropna(axis=0, how="all")
    df = df.reset_index(drop=True)

    return df


def add_current_revenue_date_columns(df, year, month_number):
    month_name = date(year, month_number, 1).strftime("%B")
    quarter = "Q" + str(((month_number - 1) // 3) + 1)
    month_date = date(year, month_number, 1)

    columns_to_remove = [
        "Year",
        "Quarter",
        "Month Name",
        "Month Number",
        "Date"
    ]

    for column in columns_to_remove:
        if column in df.columns:
            df = df.drop(columns=[column])

    df.insert(0, "Year", year)
    df.insert(1, "Quarter", quarter)
    df.insert(2, "Month Name", month_name)
    df.insert(3, "Month Number", month_number)
    df.insert(4, "Date", month_date)

    return df

def create_total_revenue_summary(df):
    zero_values_df = geography_maps.zero_value_df

    # gets all of the years in the full detailed row data
    years_in_data = df["YEAR"].dropna().unique()

    years_in_data = sorted(years_in_data)

    zero_value_dataframes = []

    for year in years_in_data:
        year = int(year)

        year_zero_values_df = zero_values_df.copy()

        year_zero_values_df["YEAR"] = year

        if "DATE" in year_zero_values_df.columns:
            year_zero_values_df["DATE"] = pd.to_datetime(
                {
                    "year": year_zero_values_df["YEAR"],
                    "month": year_zero_values_df["MONTH NUMBER"],
                    "day": 1
                }
            )

        zero_value_dataframes.append(year_zero_values_df)

    total_revenue_df = pd.concat(
        zero_value_dataframes,
        ignore_index=True
    )

    df["PAYMENT AMOUNT"] = pd.to_numeric(
        df["PAYMENT AMOUNT"],
        errors="coerce"
    ).fillna(0)

    revenue_summary = df.groupby(
        [
            "FIELD OFFICE",
            "YEAR",
            "MONTH NUMBER"
        ],
        as_index=False,
        sort=False
    )["PAYMENT AMOUNT"].sum()

    total_revenue_df = total_revenue_df.merge(
        revenue_summary,
        how="left",
        on=[
            "FIELD OFFICE",
            "YEAR",
            "MONTH NUMBER"
        ]
    )

    total_revenue_df["TOTAL REVENUE"] = total_revenue_df["PAYMENT AMOUNT"].fillna(0)

    total_revenue_df = total_revenue_df.drop(
        columns=["PAYMENT AMOUNT"],
        errors="ignore"
    )

    return total_revenue_df

def subtract_payment_amount_by_month_year_office(df1, df2):
    key_columns = [
        "YEAR",
        "MONTH NUMBER",
        "FIELD OFFICE"
    ]

    df1 = df1.copy()
    df2 = df2.copy()

    df1["TOTAL REVENUE"] = pd.to_numeric(
        df1["TOTAL REVENUE"],
        errors="coerce"
    ).fillna(0)

    df2["TOTAL REVENUE"] = pd.to_numeric(
        df2["TOTAL REVENUE"],
        errors="coerce"
    ).fillna(0)

    df2_small = df2[
        key_columns + ["TOTAL REVENUE"]
    ]

    merged_df = df1.merge(
        df2_small,
        on=key_columns,
        how="left",
        suffixes=("", "_SUBTRACT")
    )

    merged_df["TOTAL REVENUE_SUBTRACT"] = merged_df["TOTAL REVENUE_SUBTRACT"].fillna(0)

    merged_df["TOTAL REVENUE"] = (
        merged_df["TOTAL REVENUE"]
        - merged_df["TOTAL REVENUE_SUBTRACT"]
    )

    merged_df = merged_df.drop(
        columns=["TOTAL REVENUE_SUBTRACT"]
    )

    return merged_df

def save_current_revenue_workbook(total_revenue_df, community_revenue_df, grants_detail_df, output_path):
    output_path = Path(output_path)

    workbook = Workbook(write_only=True)

    total_sheet = workbook.create_sheet("Total Revenue")
    community_sheet = workbook.create_sheet("Community Revenue")
    grants_sheet = workbook.create_sheet("Grants Cleaner")

    total_df_to_save = total_revenue_df.copy()
    total_df_to_save = total_df_to_save.astype(object)
    total_df_to_save = total_df_to_save.where(pd.notna(total_df_to_save), None)

    for row in dataframe_to_rows(total_df_to_save, index=False, header=True):
        total_sheet.append(row)

    community_df_to_save = community_revenue_df.copy()
    community_df_to_save = community_df_to_save.astype(object)
    community_df_to_save = community_df_to_save.where(pd.notna(community_df_to_save), None)

    for row in dataframe_to_rows(community_df_to_save, index=False, header=True):
        community_sheet.append(row)

    grants_df_to_save = grants_detail_df.copy()
    grants_df_to_save = grants_df_to_save.astype(object)
    grants_df_to_save = grants_df_to_save.where(pd.notna(grants_df_to_save), None)

    for row in dataframe_to_rows(grants_df_to_save, index=False, header=True):
        grants_sheet.append(row)

    workbook.save(output_path)

def clean_multiple_current_revenue_imports(input_paths, input_files_names, output_path):
    output_path = Path(output_path)

    total_revenue_dataframes = []
    grants_dataframes = []

    for file_num, input_path, in enumerate(input_paths):
        original_filename = input_files_names[file_num]

        parsed_file = parse_current_revenue_filename(original_filename)
        print("About to clean", original_filename)

        if parsed_file["file_type"] == "total_revenue":
            year = parsed_file["year"]
            month_number = parsed_file["month_number"]

            cleaned_df = clean_current_revenue_dataframe(input_path)

            cleaned_df = add_current_revenue_date_columns(cleaned_df, year, month_number)

            total_revenue_dataframes.append(cleaned_df)

        elif parsed_file["file_type"] == "grants":
            cleaned_df = clean_grants_dataframe(input_path)
            grants_dataframes.append(cleaned_df)

    if len(total_revenue_dataframes) == 0:
        raise ValueError("No total revenue files were uploaded.")

    total_revenue_df = pd.concat(total_revenue_dataframes,ignore_index=True)

    # Geography Part

    total_revenue_df = total_revenue_df.rename(columns={"Contact: Mailing State/Province ↓": "State",
                                                        "Contact: Region ↑": "Field Office",
                                                        "Sum of Payment Amount": "Payment Amount"})

    # Fixes potential ffill() issue
    condition = total_revenue_df['State'].isna() & total_revenue_df['Field Office'].isna()
    total_revenue_df.loc[condition, 'State'] = "HQ"

    # Makes sure the State column is completely Capitalized
    total_revenue_df["State"] = total_revenue_df["State"].str.upper()
    total_revenue_df["State"] = total_revenue_df["State"].ffill()

    total_revenue_df = total_revenue_df.apply(geography_maps.hq_states, axis=1)

    salesforce_region_to_field_office_dict = {"Austin": "Austin Office", "Dallas": "Dallas Office",
                                              "Houston": "Houston Office", "SoCal Office": "Los Angeles Office",
                                              "SVCC Office": "Sacramento Office", "Central Florida": "Orlando Office",
                                              "West Florida": "Tampa Office", "South Florida": "Miami Office",
                                              "NOVA": "Alexandria Office", "RVA": "Richmond Office",
                                              "KCM": "Kansas City Office", "SLM": "St. Louis Office"}
    total_revenue_df["Field Office"] = total_revenue_df["Field Office"].replace(salesforce_region_to_field_office_dict)

    total_revenue_df["Field Office"] = total_revenue_df["Field Office"].fillna(total_revenue_df["State"].map(geography_maps.state_to_field_office))

    state_defaults = {
        'TX': 'Dallas Office',
        'FL': 'Miami Office',
        'CA': 'Los Angeles Office',
        'VA': 'Alexandria Office',
        'MO': 'St. Louis Office'
    }

    # Rule 1: Fill missing Field Offices using the dictionary mapping
    total_revenue_df["Field Office"] = total_revenue_df["Field Office"].fillna(total_revenue_df["State"].map(state_defaults))

    # Rule 2: Change State to "KS" if Field Office is "Kansas City Office"
    total_revenue_df.loc[total_revenue_df["Field Office"] == "Kansas City Office", "State"] = "KS"

    total_revenue_df["Region"] = total_revenue_df["State"].map(geography_maps.state_to_region)

    # Assign RSN
    total_revenue_df["Region Number"] = total_revenue_df["Region"].map(geography_maps.region_to_rsn)

    total_revenue_df["Chapter"] = total_revenue_df.apply(geography_maps.assign_chapter, axis=1)
    total_revenue_df['Region'] = total_revenue_df['Region'].str.upper()


    total_revenue_df = total_revenue_df[["Year", "Quarter", "Month Name", "Month Number", "Region", "Chapter", "State",
            "Field Office", "Region Number", "Payment Amount"]]
    total_revenue_df.columns = total_revenue_df.columns.str.upper()

    if len(grants_dataframes) > 0:
        grants_df = pd.concat(grants_dataframes, ignore_index=True)
    else:
        grants_df = pd.DataFrame()

    # changes the name of some columns

    grants_df = grants_df.rename(columns={'Payment Date': 'Date'})

    # Changes ZIP Code column to string, fixes leading zeroes, and takes only the first 5 digits
    grants_df["Zipcode"] = grants_df["Zipcode"].astype(str).str.extract(r'(\d{1,5})')[0].str.zfill(5)

    # Makes sure the State column is completely Capitalized
    grants_df["State"] = grants_df["State"].str.upper()

    grants_df = grants_df.apply(geography_maps.hq_states, axis=1)

    grants_df["Region"] = grants_df["State"].map(geography_maps.state_to_region)

    # Assign RSN
    grants_df["Region Number"] = grants_df["Region"].map(geography_maps.region_to_rsn)

    # Assign Field Office
    grants_df["Field Office"] = grants_df["State"].map(geography_maps.state_to_field_office)

    grants_df = grants_df.apply(lambda row: geography_maps.assign_field_office(row, True),axis=1)

    grants_df["Chapter"] = grants_df.apply(geography_maps.assign_chapter, axis=1)
    grants_df['Region'] = grants_df['Region'].str.upper()

    grants_df = geography_maps._add_date_columns(grants_df)

    if len(grants_dataframes) > 0:
        grants_detail_df = build_grants_detail_worksheet(grants_df)
    else:
        grants_detail_df = pd.DataFrame(columns=[
            "REGION", "RSN", "CHAPTER", "Contact State ↓", "FILED OFFICE",
            "Donation Record Type", "Payment Date", "Contact: Full Name",
            "Payment Amount", "Campaign Name", "GOAL AMOUNT", "City",
            "Payment Method", "ICNA Donations Types", "Notes", "Lead Source",
            "Donation ID", "Comments", "Donation Ref", "Donations Snapshot ID",
            "DONATION NAME", "Year", "Month Name", "Month Number", "PSN",
            "QUARTER",
        ])

    grants_df.columns = grants_df.columns.str.upper()

    grants_summary_df = create_total_revenue_summary(grants_df)

    total_revenue_summary_df = create_total_revenue_summary(total_revenue_df)

    community_revenue_summary_df = subtract_payment_amount_by_month_year_office(total_revenue_summary_df, grants_summary_df)

    save_current_revenue_workbook(total_revenue_summary_df, community_revenue_summary_df, grants_detail_df, output_path)