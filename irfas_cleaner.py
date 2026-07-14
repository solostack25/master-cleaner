from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
import geography_maps
import re


REQUIRED_HEADER_COLUMNS = ["Case ID", "Service Requested"]


def clean_text(value):
    if value is None:
        return ""

    return str(value).strip().lower()


def unmerge_first_sheet(input_path):
    input_path = Path(input_path)

    workbook = load_workbook(input_path)
    sheet = workbook.worksheets[0]

    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        top_left_cell = sheet.cell(
            row=merged_range.min_row,
            column=merged_range.min_col
        )

        top_left_value = top_left_cell.value

        sheet.unmerge_cells(str(merged_range))

        for row_number in range(merged_range.min_row, merged_range.max_row + 1):
            for column_number in range(merged_range.min_col, merged_range.max_col + 1):
                sheet.cell(
                    row=row_number,
                    column=column_number
                ).value = top_left_value

    temporary_path = input_path.parent / f"{input_path.stem}_irfas_unmerged_temp.xlsx"

    workbook.save(temporary_path)
    workbook.close()

    return temporary_path


def find_header_row(input_path):
    input_path = Path(input_path)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    required_headers = {
        clean_text(column)
        for column in REQUIRED_HEADER_COLUMNS
    }

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        row_values = {
            clean_text(cell)
            for cell in row
        }

        if required_headers.issubset(row_values):
            workbook.close()

            # pandas uses zero-based row numbers for the header argument.
            return row_number - 1

    workbook.close()

    raise ValueError(
        "Could not find the IRFAS header row. "
        "Expected a row containing Case ID and Service Requested."
    )

def irfas_zero_value_rows(df):
    new_dfs = []

    programs = [
        "Hunger Prevention",
        "General",
        "Refugee Services & Community Empowerment",
        "Transitional Housing",
        "Muslim Family Services",
        "Health Services",
        "Back 2 School",
        "FATE",
        "Disaster Relief"
    ]

    # Copy the 732-row zero value geography/month dataframe.
    base_df = geography_maps.zero_value_df.copy()
    base_df.columns = base_df.columns.map(lambda col: str(col).strip().upper())

    # Get years that already exist in the data.
    years = sorted(df["YEAR"].dropna().astype(int).unique())

    for year in years:
        for program in programs:
            temp_df = base_df.copy()

            # Use the year from the actual uploaded data.
            temp_df["YEAR"] = year

            # If you want the DATE year to match the current year,
            # keep this part. It uses the month already in base_df.
            temp_df["DATE"] = pd.to_datetime(temp_df["DATE"], errors="coerce")
            temp_df["DATE"] = temp_df["DATE"].apply(
                lambda date_value: pd.Timestamp(
                    year=year,
                    month=date_value.month,
                    day=1
                ).date()
                if pd.notna(date_value)
                else date_value
            )

            # Set the program.
            temp_df["PROGRAM"] = program

            # Make sure temp_df has every column that df has.
            for column in df.columns:
                if column not in temp_df.columns:
                    temp_df[column] = ""

            # Keep the same column order as df.
            temp_df = temp_df[df.columns]

            new_dfs.append(temp_df)

    if new_dfs:
        new_df = pd.concat(new_dfs, ignore_index=True)
        df = pd.concat([df, new_df], ignore_index=True)

    return df

def clean_irfas(input_path, output_path):
    input_path = Path(input_path)
    output_path = Path(output_path)

    temporary_path = unmerge_first_sheet(input_path)

    try:
        header_row_index = find_header_row(temporary_path)

        df = pd.read_excel(
            temporary_path,
            sheet_name=0,
            header=header_row_index
        )

        df = df.dropna(axis=0, how="all")
        df = df.dropna(axis=1, how="all")

        df.columns = df.columns.map(lambda column: str(column).strip())


    finally:
        try:
            temporary_path.unlink()
        except Exception:
            pass

    # Assigns the status to every row.
    row_status_list = []

    current_status = "Approved"
    list_of_status = ["Approved", "Void", "Declined"]

    for index, row in df.iterrows():
        case_id = row["Case ID"]

        if pd.isna(case_id):
            case_id_text = ""
        else:
            case_id_text = str(case_id).strip()

        if case_id_text in list_of_status:
            current_status = case_id_text

        row_status_list.append(current_status)

    df["Status"] = row_status_list

    # Removes the rows that are column headers or headings for status.
    rows_to_delete = []

    for index, row in df.iterrows():
        case_id = row["Case ID"]

        if pd.isna(case_id):
            case_id_text = ""
        else:
            case_id_text = str(case_id).strip()

        if case_id_text in list_of_status:
            rows_to_delete.append(index)

        if case_id_text == "Case ID":
            rows_to_delete.append(index)

    df = df.drop(index=rows_to_delete).reset_index(drop=True)

    # Geography Assigning
    df["State"] = df["State"].str.upper()
    df.rename(columns={'Zip Code': 'Zipcode', 'Decision Date': 'Date'}, inplace=True)

    df = df.apply(geography_maps.hq_states, axis=1)

    df["Region"] = df["State"].map(geography_maps.state_to_region)

    # Assign RSN
    df["Region Number"] = df["Region"].map(geography_maps.region_to_rsn)

    # Assign Field Office
    df["Field Office"] = df["State"].map(geography_maps.state_to_field_office)

    # Changes ZIP Code column to string, fixes leading zeroes, and takes only the first 5 digits
    df["Zipcode"] = df["Zipcode"].astype(str).str.extract(r'(\d{1,5})')[0].str.zfill(5)

    df = df.apply(
        lambda row: geography_maps.assign_field_office(row, True),
        axis=1
    )

    df["Chapter"] = df.apply(geography_maps.assign_chapter, axis=1)
    df['Region'] = df['Region'].str.upper()

    # Date Stuff

    df = geography_maps._add_date_columns(df)
    df.rename(columns={'Date': 'Decision Date', 'Zipcode': 'Zip Code'}, inplace=True)

    # Program Assignment

    # Default program
    df["Program"] = "General"

    for keyword, program in geography_maps.mapping.items():
        pattern = r"\b{}\b".format(re.escape(keyword))

        # Check Service Requested
        mask_name = df["Service Requested"].str.contains(pattern, case=False, na=False, regex=True)
        df.loc[(df["Program"] == "General") & mask_name, "Program"] = program


    df["Count of Application"] = 1

    new_columns = ["Revenue Raised Total", "Community Revenue Raised", "Grants Revenue Raised",
                   "Financial Assistance via Community Revenue", "Financial Assistance via Grants", "Limits"]

    df[new_columns] = 0

    df = df[["Year", "Quarter", "Month Name", "Month Number", "Decision Date", "Region", "Chapter", "State",
            "Field Office", "Region Number", "Case ID", "City", "Zip Code", "Check Date", "Check #", "Amount Approved",
             "Approved By", "Recommended By", "Service Requested", "Program", "Grant Name", "Revenue Raised Total",
             "Community Revenue Raised", "Grants Revenue Raised", "Financial Assistance via Community Revenue",
             "Financial Assistance via Grants", "Limits"]]




    # Capitalizes All Columns
    df.columns = df.columns.str.upper()

    # adds zero value rows
    df = irfas_zero_value_rows(df)


    df.to_excel(output_path, index=False)

    # Currency Formatting
    currency_columns = new_columns + ["Amount Approved"]
    for column in currency_columns:
        geography_maps.format_currency_column(output_path, column.upper())
