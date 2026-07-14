from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from datetime import date

import geography_maps

EXPENDITURE_START_VALUES = ["Admin", "Finance"]

dispersion_df = pd.read_csv("Office Dispersion Logic.csv")

def clean_text(value):
    if pd.isna(value):
        return ""

    return str(value).strip()


def save_expenditure_workbook(total_df, grant_df, central_expenses_df, output_path):
    workbook = Workbook(write_only=True)

    total_sheet = workbook.create_sheet("Total Expenses")
    grant_sheet = workbook.create_sheet("Grant Expenses")
    central_sheet = workbook.create_sheet("Central Expenses")

    total_df_to_save = total_df.copy()
    total_df_to_save = total_df_to_save.astype(object)
    office_df_to_save = total_df_to_save.where(pd.notna(total_df_to_save), None)

    for row in dataframe_to_rows(office_df_to_save, index=False, header=True):
        total_sheet.append(row)

    grant_df_to_save = grant_df.copy()
    grant_df_to_save = grant_df_to_save.astype(object)
    grant_df_to_save = grant_df_to_save.where(pd.notna(grant_df_to_save), None)

    for row in dataframe_to_rows(grant_df_to_save, index=False, header=True):
        grant_sheet.append(row)

    central_df_to_save = central_expenses_df.copy()
    central_df_to_save = central_df_to_save.astype(object)
    central_df_to_save = central_df_to_save.where(pd.notna(central_df_to_save), None)

    for row in dataframe_to_rows(central_df_to_save, index=False, header=True):
        central_sheet.append(row)

    workbook.save(output_path)


def find_expenditure_start_row(df):
    start_values = []

    for value in EXPENDITURE_START_VALUES:
        start_values.append(value.strip().lower())

    for row_index, row in df.iterrows():
        for value in row.values:
            value_text = clean_text(value).lower()

            if value_text in start_values:
                return row_index

    raise ValueError(
        "Could not find the expenditure starting row. "
        "Expected to find a row containing Admin or Finance."
    )


def demerge_expenditure(input_path):
    input_path = Path(input_path)

    workbook = load_workbook(input_path, data_only=True)
    sheet = workbook.worksheets[0]

    merged_ranges = list(sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        top_left_cell = sheet.cell(
            row=merged_range.min_row,
            column=merged_range.min_col
        )

        top_left_value = top_left_cell.value

        sheet.unmerge_cells(str(merged_range))

        for row_number in range(merged_range.min_row, merged_range.max_row + 1):
            for column_number in range(merged_range.min_col, merged_range.max_col + 1):
                sheet.cell(
                    row=row_number,
                    column=column_number
                ).value = top_left_value

    rows = []

    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))

    workbook.close()

    df = pd.DataFrame(rows)

    # Treat blank-looking cells as actually empty.
    df = df.replace(r"^\s*$", pd.NA, regex=True)

    # Delete rows and columns that are completely empty.
    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")

    df = df.reset_index(drop=True)

    # Find the first real row of the expenditure data.
    start_row_index = find_expenditure_start_row(df)

    # Delete everything above that row.
    df = df.iloc[start_row_index:].copy()
    df = df.reset_index(drop=True)

    # Clean empty rows and columns again after deleting the top section.
    df = df.replace(r"^\s*$", pd.NA, regex=True)
    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")

    df = df.reset_index(drop=True)

    return df

def get_first_non_empty_column_number(df):
    for column_number, column_name in enumerate(df.columns):
        # Convert to string and strip whitespace to catch "   " or ""
        name_str = str(column_name).strip()

        # Check against common empty headers: NaN, 'nan', or empty string
        if not pd.isna(column_name) and name_str != '' and name_str.lower() != 'nan':
            return column_number

    return None

def transpose_expenditure(df):
    output_df = df.set_index('Expense Category L2').T.reset_index()

    # Clear the name of the index axis
    output_df.columns.name = "Location"

    return output_df

def normalize_office_text(value):
    if pd.isna(value):
        return ""

    value = str(value).strip().lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip()

    return value


def prepare_dispersion_df(dispersion_df):
    dispersion_df = dispersion_df.copy()

    dispersion_df.columns = dispersion_df.columns.str.strip().str.upper()

    required_columns = [
        "REGION",
        "REGION NUMBER",
        "CHAPTER",
        "STATE",
        "FIELD OFFICE",
        "FTE"
    ]

    for column in required_columns:
        if column not in dispersion_df.columns:
            raise ValueError(f"Missing required dispersion column: {column}")

    text_columns = [
        "REGION",
        "CHAPTER",
        "STATE",
        "FIELD OFFICE"
    ]

    for column in text_columns:
        dispersion_df[column] = dispersion_df[column].astype(str).str.strip()

    dispersion_df["STATE"] = dispersion_df["STATE"].str.upper()
    dispersion_df["FTE"] = pd.to_numeric(dispersion_df["FTE"], errors="coerce")

    dispersion_df = dispersion_df.reset_index(drop=True)

    return dispersion_df


def build_office_lookup(dispersion_df):
    office_lookup = {}

    for idx, row in dispersion_df.iterrows():
        field_office = row["FIELD OFFICE"]

        normalized_field_office = normalize_office_text(field_office)

        office_lookup[normalized_field_office] = row

        if field_office.endswith(" Office"):
            shorter_name = field_office.replace(" Office", "").strip()
            office_lookup[normalize_office_text(shorter_name)] = row

        if field_office.endswith("-Statewide"):
            shorter_name = field_office.replace("-Statewide", "").strip()
            office_lookup[normalize_office_text(shorter_name)] = row

    return office_lookup

def split_state_divide_rows(df, dispersion_df):
    dispersion_df = prepare_dispersion_df(dispersion_df)

    expense_column = "Expense"

    amount_columns = []

    for column in df.columns:
        if column != expense_column:
            amount_columns.append(column)

    output_rows = []

    for idx, row in df.iterrows():
        expense_value = row[expense_column]

        if pd.isna(expense_value):
            output_rows.append(row)
            continue

        expense_text = str(expense_value).strip().upper()

        match = re.match(r"^([A-Z]{2})\s+DIVIDE$", expense_text)

        if match is None:
            output_rows.append(row)
            continue

        state = match.group(1)

        if state == "MO":
            states_to_use = ["MO", "KS"]
        else:
            states_to_use = [state]

        state_dispersion_df = dispersion_df[
            dispersion_df["STATE"].isin(states_to_use)
        ]

        if state_dispersion_df.empty:
            raise ValueError(f"No dispersion rows found for state {state}.")

        total_fte = state_dispersion_df["FTE"].sum()

        if pd.isna(total_fte) or total_fte == 0:
            raise ValueError(f"Total FTE for state {state} is missing or zero.")

        for dispersion_idx, dispersion_row in state_dispersion_df.iterrows():
            field_office = dispersion_row["FIELD OFFICE"]
            fte = dispersion_row["FTE"]

            if pd.isna(fte) or fte == 0:
                continue

            ratio = fte / total_fte

            new_row = row.copy()
            new_row[expense_column] = field_office

            for column in amount_columns:
                amount_value = pd.to_numeric(new_row[column], errors="coerce")

                if pd.isna(amount_value):
                    new_row[column] = pd.NA
                else:
                    new_row[column] = amount_value * ratio

            output_rows.append(new_row)

    output_df = pd.DataFrame(output_rows)
    output_df = output_df.reset_index(drop=True)

    return output_df

def remove_grants_from_expense(expense_text):
    expense_text = str(expense_text).strip()

    if expense_text.lower().endswith("grants"):
        return expense_text[:-6].strip()

    return expense_text


def add_expense_type_field_office_geography(df, dispersion_df):
    dispersion_df = prepare_dispersion_df(dispersion_df)
    office_lookup = build_office_lookup(dispersion_df)

    columns_to_remove = [
        "Expense Type",
        "Region",
        "Region Number",
        "Chapter",
        "State",
        "Field Office"
    ]

    for column in columns_to_remove:
        if column in df.columns:
            df = df.drop(columns=[column])

    expense_type_list = []
    region_list = []
    region_number_list = []
    chapter_list = []
    state_list = []
    field_office_list = []

    for idx, row in df.iterrows():
        expense = row["Expense"]

        if pd.isna(expense):
            expense_text = ""
        else:
            expense_text = str(expense).strip()

        if expense_text == "":
            expense_type = pd.NA
            geography_row = None

        elif expense_text.lower().startswith("central"):
            expense_type = "Central"

            normalized_office_name = normalize_office_text("Unassigned")
            geography_row = office_lookup[normalized_office_name]

        elif expense_text.lower().endswith("grants"):
            expense_type = "Grant"

            office_name = remove_grants_from_expense(expense_text)
            normalized_office_name = normalize_office_text(office_name)

            if normalized_office_name not in office_lookup:
                raise ValueError(
                    f"Could not match grant expense to field office: {expense_text}"
                )

            geography_row = office_lookup[normalized_office_name]

        else:
            expense_type = "Office"

            normalized_expense = normalize_office_text(expense_text)

            if normalized_expense not in office_lookup:
                raise ValueError(
                    f"Could not match office expense to field office: {expense_text}"
                )

            geography_row = office_lookup[normalized_expense]

        expense_type_list.append(expense_type)

        if geography_row is None:
            region_list.append(pd.NA)
            region_number_list.append(pd.NA)
            chapter_list.append(pd.NA)
            state_list.append(pd.NA)
            field_office_list.append(pd.NA)

        else:
            region_list.append(geography_row["REGION"])
            region_number_list.append(geography_row["REGION NUMBER"])
            chapter_list.append(geography_row["CHAPTER"])
            state_list.append(geography_row["STATE"])
            field_office_list.append(geography_row["FIELD OFFICE"])

    expense_column_number = df.columns.get_loc("Expense")

    df.insert(expense_column_number + 1, "Expense Type", expense_type_list)
    df.insert(expense_column_number + 2, "Region", region_list)
    df.insert(expense_column_number + 3, "Region Number", region_number_list)
    df.insert(expense_column_number + 4, "Chapter", chapter_list)
    df.insert(expense_column_number + 5, "State", state_list)
    df.insert(expense_column_number + 6, "Field Office", field_office_list)

    return df

def get_grant_expense_name(field_office):
    field_office = str(field_office).strip()

    if field_office.endswith(" Office"):
        base_name = field_office.replace(" Office", "").strip()

    elif field_office.endswith("-Statewide"):
        base_name = field_office.replace("-Statewide", "").strip()

    else:
        base_name = field_office

    return base_name + " Grants"


def add_expenditure_zero_value_rows(df, dispersion_df):
    dispersion_df = prepare_dispersion_df(dispersion_df)

    id_columns = [
        "Expense",
        "Expense Type",
        "Region",
        "Region Number",
        "Chapter",
        "State",
        "Field Office"
    ]

    amount_columns = []

    for column in df.columns:
        if column not in id_columns:
            amount_columns.append(column)

    zero_rows = []

    for idx, office_row in dispersion_df.iterrows():
        field_office = office_row["FIELD OFFICE"]

        office_zero_row = {}

        office_zero_row["Expense"] = field_office
        office_zero_row["Expense Type"] = "Office"
        office_zero_row["Region"] = office_row["REGION"]
        office_zero_row["Region Number"] = office_row["REGION NUMBER"]
        office_zero_row["Chapter"] = office_row["CHAPTER"]
        office_zero_row["State"] = office_row["STATE"]
        office_zero_row["Field Office"] = field_office

        for column in amount_columns:
            office_zero_row[column] = 0

        zero_rows.append(office_zero_row)

        grant_zero_row = {}

        grant_zero_row["Expense"] = get_grant_expense_name(field_office)
        grant_zero_row["Expense Type"] = "Grant"
        grant_zero_row["Region"] = office_row["REGION"]
        grant_zero_row["Region Number"] = office_row["REGION NUMBER"]
        grant_zero_row["Chapter"] = office_row["CHAPTER"]
        grant_zero_row["State"] = office_row["STATE"]
        grant_zero_row["Field Office"] = field_office

        for column in amount_columns:
            grant_zero_row[column] = 0

        zero_rows.append(grant_zero_row)

    zero_df = pd.DataFrame(zero_rows)

    df = pd.concat(
        [
            df,
            zero_df
        ],
        ignore_index=True
    )

    return df

def combine_duplicate_expenditure_rows(df):
    group_columns = [
        "Expense",
        "Expense Type",
        "Region",
        "Region Number",
        "Chapter",
        "State",
        "Field Office"
    ]

    amount_columns = []

    for column in df.columns:
        if column not in group_columns:
            amount_columns.append(column)

    for column in amount_columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    df = df.groupby(
        group_columns,
        as_index=False,
        sort=False,
        dropna=False
    )[amount_columns].sum(min_count=1)

    return df

def add_report_month_columns(df, report_month):
    year = int(report_month[0:4])
    month_number = int(report_month[5:7])

    month_name = date(year, month_number, 1).strftime("%B")
    quarter = "Q" + str(((month_number - 1) // 3) + 1)

    df.insert(0, "Year", year)
    df.insert(1, "Quarter", quarter)
    df.insert(2, "Month Name", month_name)
    df.insert(3, "Month Number", month_number)

    return df

def allocate_central_expenses(df, dispersion_df):
    dispersion_df = prepare_dispersion_df(dispersion_df)

    central_expenses_df = df[
        df["Expense Type"] == "Central"
    ].copy()

    id_columns = [
        "Expense",
        "Expense Type",
        "Region",
        "Region Number",
        "Chapter",
        "State",
        "Field Office"
    ]

    amount_columns = []

    for column in df.columns:
        if column not in id_columns:
            amount_columns.append(column)

    central_total = 0

    for column in amount_columns:
        central_column_total = pd.to_numeric(
            central_expenses_df[column],
            errors="coerce"
        ).sum()

        central_total = central_total + central_column_total

    total_fte = dispersion_df["FTE"].sum()

    if pd.isna(total_fte) or total_fte == 0:
        raise ValueError("Total national FTE is missing or zero.")

    fte_lookup = {}

    for idx, row in dispersion_df.iterrows():
        field_office = row["FIELD OFFICE"]
        fte = row["FTE"]

        normalized_field_office = normalize_office_text(field_office)

        fte_lookup[normalized_field_office] = fte

    central_expenses_list = []

    for idx, row in df.iterrows():
        expense_type = row["Expense Type"]
        field_office = row["Field Office"]

        central_expense_amount = 0

        # I am assigning central expenses only to Office rows.
        # Do not assign to Grant rows, or central expenses would be duplicated.
        if expense_type == "Office":
            normalized_field_office = normalize_office_text(field_office)

            if normalized_field_office in fte_lookup:
                office_fte = fte_lookup[normalized_field_office]

                if not pd.isna(office_fte):
                    office_ratio = office_fte / total_fte
                    central_expense_amount = central_total * office_ratio

        central_expenses_list.append(central_expense_amount)

    df["Central Expenses"] = central_expenses_list

    df = df[
        df["Expense Type"] != "Central"
    ]

    df = df.reset_index(drop=True)

    return df, central_expenses_df

def reorder_and_uppercase_expenditure_columns(df):
    first_columns = [
        "Year",
        "Quarter",
        "Month Name",
        "Month Number",
        "Region",
        "Chapter",
        "State",
        "Field Office",
        "Region Number"
    ]

    existing_first_columns = []

    for column in first_columns:
        if column in df.columns:
            existing_first_columns.append(column)

    remaining_columns = []

    for column in df.columns:
        if column not in existing_first_columns:
            remaining_columns.append(column)

    df = df[existing_first_columns + remaining_columns]

    df.columns = df.columns.str.upper()

    return df

def combine_main_df_by_field_office(df):
    columns_to_remove = [
        "Expense",
        "Expense Type"
    ]

    group_columns = [
        "Year",
        "Quarter",
        "Month Name",
        "Month Number",
        "Region",
        "Region Number",
        "Chapter",
        "State",
        "Field Office"
    ]

    df = df.drop(columns=columns_to_remove)

    amount_columns = []

    for column in df.columns:
        if column not in group_columns:
            amount_columns.append(column)

    for column in amount_columns:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    df = df.groupby(
        group_columns,
        as_index=False,
        sort=False,
        dropna=False
    )[amount_columns].sum(min_count=1)

    return df

def clean_expenditure(input_path, output_path, report_month):
    input_path = Path(input_path)
    output_path = Path(output_path)

    df = demerge_expenditure(input_path)

    # 1. Set the 3rd row (index 2) as the new header
    df.columns = df.iloc[1]

    # 2. Slice the dataframe to keep everything AFTER the 3rd row (index 3 onwards)
    df = df.iloc[2:]

    # 3. Clean up the index and clear any residual header name
    df.reset_index(drop=True, inplace=True)
    df.columns.name = None

    # Drop the first row from the data and reset the index
    df = df[1:].reset_index(drop=True)

    # Clean up the header name attribute if necessary
    df.columns.name = None

    first_non_empty_column_number = get_first_non_empty_column_number(df)

    if first_non_empty_column_number is None:
        raise ValueError("Could not find any non-empty column.")

    if first_non_empty_column_number < 3:
        print("Number of non-empty columns is", first_non_empty_column_number)
        print(list(df.columns))
        raise ValueError("There are not 3 columns before the first non-empty column.")

    df = df.iloc[:, first_non_empty_column_number - 3:]
    df = df.reset_index(drop=True)

    columns = list(df.columns)

    # After slicing, these are now the first 3 columns
    columns[0] = "Expense Category L1"
    columns[1] = "Expense Category L2"
    columns[2] = "Expense Category L3"

    df.columns = columns

    # Deletes all rows where L1 and L2 start with "Total"
    df["Expense Category L1"] = df["Expense Category L1"].astype("string")
    df["Expense Category L2"] = df["Expense Category L2"].astype("string")

    df = df[
        ~df["Expense Category L1"].str.strip().str.startswith("Total", na=False)
    ]
    df = df[
        ~df["Expense Category L2"].str.strip().str.startswith("Total", na=False)
    ]

    df = df.reset_index(drop=True)

    df["Expense Category L1"] = df["Expense Category L1"].ffill()

    # Delete all rows where Expense Category Levels 2 and 3 are empty

    df = df.replace(r"^\s*$", pd.NA, regex=True)

    df = df.dropna(
        subset=[
            "Expense Category L2",
            "Expense Category L3"
        ],
        how="all"
    )

    df = df.reset_index(drop=True)


    # Deletes unnecessary rows. We find rows where L3 is empty before an L3 that is full and deletes the previous row
    rows_to_delete = []

    df = df.replace(r"^\s*$", pd.NA, regex=True)
    df = df.reset_index(drop=True)

    for idx, row in df.iterrows():
        if idx == 0:
            continue

        current_l3 = row["Expense Category L3"]
        previous_l3 = df["Expense Category L3"].iloc[idx - 1]

        if not pd.isna(current_l3):
            df.loc[idx, "Expense Category L2"] = current_l3
            if pd.isna(previous_l3):
                rows_to_delete.append(idx - 1)

    df = df.drop(index=rows_to_delete)
    df = df.reset_index(drop=True)

    df = df.drop(columns=["Expense Category L3"])
    df = df.drop(columns=["Expense Category L1"])
    df = df.reset_index(drop=True)

    total_column_number = None

    for column_number, column_name in enumerate(df.columns):
        column_name_text = str(column_name).strip().upper()

        if column_name_text == "TOTAL":
            total_column_number = column_number
            break

    if total_column_number is not None:
        df = df.iloc[:, :total_column_number]

    # Transpose the table.
    df = transpose_expenditure(df)

    columns = list(df.columns)
    columns[0] = "Expense"
    df.columns = columns

    # Split rows like "TX Divide", "CA Divide", "MO Divide"
    df = split_state_divide_rows(df, dispersion_df)

    # Add Expense Type, Region, Region Number, Chapter, State, Field Office
    df = add_expense_type_field_office_geography(df, dispersion_df)

    # Add zero-value Office and Grant rows for every office
    df = add_expenditure_zero_value_rows(df, dispersion_df)

    # Combine duplicate rows now that zero rows are included
    df = combine_duplicate_expenditure_rows(df)

    # Pull central expenses into separate dataframe,
    # allocate central total back to office rows,
    # and remove central rows from main df
    df, central_expenses_df = allocate_central_expenses(df, dispersion_df)

    # Add selected month/year from website
    df = add_report_month_columns(df, report_month)
    central_expenses_df = add_report_month_columns(central_expenses_df, report_month)

    # Create grant sheet before transforming df into field-office-level main df
    grant_df = df[df["Expense Type"] == "Grant"]
    grant_df = grant_df.reset_index(drop=True)

    # Main df becomes field-office-level total.
    # This includes Office + Grant rows added together.
    # Expense and Expense Type go away here.
    df = combine_main_df_by_field_office(df)

    # Reorder and uppercase each dataframe based on its own columns
    df = reorder_and_uppercase_expenditure_columns(df)
    grant_df = reorder_and_uppercase_expenditure_columns(grant_df)
    central_expenses_df = reorder_and_uppercase_expenditure_columns(central_expenses_df)

    # Save final output
    save_expenditure_workbook(
        df,
        grant_df,
        central_expenses_df,
        output_path
    )